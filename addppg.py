from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from ppg import getStats

wb = load_workbook('PPG.xlsx')
ws = wb.active

i = 0

for col in range(2, 33):
    char = get_column_letter(col)
    ws[char + str(3)] = getStats()[i]
    i += 1

wb.save(filename="PPG.xlsx")
