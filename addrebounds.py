from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from rebounds import getStats

wb = load_workbook('Rebounds.xlsx')
ws = wb.active

i = 0

for col in range(2, 17):
    char = get_column_letter(col)
    for x in range(2, 7):
        ws[char + str(x)] = getStats()[i]
        i += 1

wb.save(filename="Rebounds.xlsx")
