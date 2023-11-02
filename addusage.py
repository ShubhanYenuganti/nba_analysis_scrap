from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from usage import getStats

wb = load_workbook('Usage_Rate.xlsx')
ws = wb.active

i = 0

for col in range(2, 33):
    char = get_column_letter(col)
    for x in range(3, 8):
        ws[char + str(x)] = getStats()[i]
        i += 1

wb.save(filename="Usage_Rate.xlsx")
